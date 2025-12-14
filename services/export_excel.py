# -*- coding: utf-8 -*-
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime


def build_excel_report(history: list) -> BytesIO:
    """
    Собирает Excel-отчёт по сохранённой истории пользователя.
    FSM и Telegram сюда НЕ заходят.
    """
    wb = Workbook()
    ws_map = {}

    for item in history:
        t = item.get("type", "other")

        if t not in ws_map:
            ws = wb.create_sheet(title=t)
            ws.append(["Дата", "Сводка", "Данные"])
            ws_map[t] = ws

        ws = ws_map[t]
        ws.append([
            item.get("date", ""),
            item.get("summary", ""),
            str(item.get("raw_data", {})),
        ])

    # удалить пустой дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
